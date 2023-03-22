
import openpyxl, random
import pandas as pd
from openpyxl import Workbook

rbook = openpyxl.load_workbook('dataset.xlsx', data_only=True)

rsheet = rbook.active

wbook = Workbook()
wsheet = wbook.active

print(rsheet.cell(row=2, column=2).value)

wsheet.cell(row=1, column=1).value = "Student ID"
wsheet.cell(row=1, column=2).value = "Academics"
wsheet.cell(row=1, column=3).value = "Looks_Fitness"
wsheet.cell(row=1, column=4).value = "Social_life"
wsheet.cell(row=1, column=5).value = "Xtra_curricular"
wsheet.cell(row=1, column=6).value = "Athletics"
wsheet.cell(row=1, column=7).value = "Career"
wsheet.cell(row=1, column=8).value = "Finance"
wsheet.cell(row=1, column=9).value = "Relationship"
wsheet.cell(row=1, column=10).value = "Cultural_Shock"
wsheet.cell(row=1, column=11).value = "Emotional_bullied"
wsheet.cell(row=1, column=12).value = "Physical_bullied"
wsheet.cell(row=1, column=13).value = "Verbal_bullied"
wsheet.cell(row=1, column=14).value = "Social_bullied"
wsheet.cell(row=1, column=15).value = "Cyber_bullied"
wsheet.cell(row=1, column=16).value = "International"
wsheet.cell(row=1, column=17).value = "Miss_home"
wsheet.cell(row=1, column=18).value = "Family_friends"
wsheet.cell(row=1, column=19).value = "Food"
wsheet.cell(row=1, column=20).value = "Sensory"
wsheet.cell(row=1, column=21).value = "Miss_social"
wsheet.cell(row=1, column=22).value = "Native_language"
wsheet.cell(row=1, column=23).value = "Courses"
wsheet.cell(row=1, column=24).value = "Loan"
wsheet.cell(row=1, column=25).value = "Stressed_commute"
wsheet.cell(row=1, column=26).value = "Stress_rating"


count = 0
data = rsheet.cell(row=count+1, column=1).value
while data != None:
    count += 1
    data = rsheet.cell(row=count+1, column=1).value


for i in range(2, count+1):
    wsheet.cell(row=i, column=1).value = i - 1

    data = rsheet.cell(row=i, column=2).value
    if "Academic" in data:
        wsheet.cell(row=i, column=2).value = 1
    else:
        wsheet.cell(row=i, column=2).value = 0
    if "fitness" in data:
        wsheet.cell(row=i, column=3).value = 1
    else:
        wsheet.cell(row=i, column=3).value = 0
    if "Social" in data:
        wsheet.cell(row=i, column=4).value = 1
    else:
        wsheet.cell(row=i, column=4).value = 0
    if "extra-curricular" in data:
        wsheet.cell(row=i, column=5).value = 1
    else:
        wsheet.cell(row=i, column=5).value = 0
    if "Athletic" in data:
        wsheet.cell(row=i, column=6).value = 1
    else:
        wsheet.cell(row=i, column=6).value = 0
    if "Career" in data:
        wsheet.cell(row=i, column=7).value = 1
    else:
        wsheet.cell(row=i, column=7).value = 0
    if "Financial" in data:
        wsheet.cell(row=i, column=8).value = 1
    else:
        wsheet.cell(row=i, column=8).value = 0
    if "Relationship" in data:
        wsheet.cell(row=i, column=9).value = 1
    else:
        wsheet.cell(row=i, column=9).value = 0
    if "Cultural" in data:
        wsheet.cell(row=i, column=10).value = 1
    else:
        wsheet.cell(row=i, column=10).value = 0

    data = rsheet.cell(row=i, column=3).value
    if "Mentally/emotionally bullied" in data:
        wsheet.cell(row=i, column=11).value = 1
    else:
        wsheet.cell(row=i, column=11).value = 0
    if "Physically bullied" in data:
        wsheet.cell(row=i, column=12).value = 1
    else:
        wsheet.cell(row=i, column=12).value = 0
    if "Verbal bullying" in data:
        wsheet.cell(row=i, column=13).value = 1
    else:
        wsheet.cell(row=i, column=13).value = 0
    if "Social bullying" in data:
        wsheet.cell(row=i, column=14).value = 1
    else:
        wsheet.cell(row=i, column=14).value = 0
    if "Cyber bullying" in data:
        wsheet.cell(row=i, column=15).value = 1
    else:
        wsheet.cell(row=i, column=15).value = 0

    data = rsheet.cell(row=i, column=4).value
    if "Yes" in data:
        wsheet.cell(row=i, column=16).value = 1
    else:
        wsheet.cell(row=i, column=16).value = 0

    data = rsheet.cell(row=i, column=5).value
    if "No, I don't miss my home" in data:
        wsheet.cell(row=i, column=17).value = 0
    else:
        wsheet.cell(row=i, column=17).value = 1
    if "Yes, Family and friends" in data:
        wsheet.cell(row=i, column=18).value = 1
    else:
        wsheet.cell(row=i, column=18).value = 0
    if "Yes, Food" in data:
        wsheet.cell(row=i, column=19).value = 1
    else:
        wsheet.cell(row=i, column=19).value = 0
    if "Yes, Sensory experience of staying in home" in data:
        wsheet.cell(row=i, column=20).value = 1
    else:
        wsheet.cell(row=i, column=20).value = 0
    if "Yes, Social life of my hometown" in data:
        wsheet.cell(row=i, column=21).value = 1
    else:
        wsheet.cell(row=i, column=21).value = 0
    if "Yes, Native language conversations" in data:
        wsheet.cell(row=i, column=22).value = 1
    else:
        wsheet.cell(row=i, column=22).value = 0
    
    data = rsheet.cell(row=i, column=6).value
    if data == 1:
        wsheet.cell(row=i, column=23).value = 1
    elif data == 2:
        wsheet.cell(row=i, column=23).value = 2
    elif data == 3:
        wsheet.cell(row=i, column=23).value = 3
    elif data == 4:
        wsheet.cell(row=i, column=23).value = 4
    elif data == "More than 4":
        wsheet.cell(row=i, column=23).value = 5
    else:
        wsheet.cell(row=i, column=23).value = 0

    data = rsheet.cell(row=i, column=7).value
    if "Loan" in data:
        wsheet.cell(row=i, column=24).value = 1
    else:
        wsheet.cell(row=i, column=24).value = 0

    data = rsheet.cell(row=i, column=8).value
    if "Yes" in data or "Maybe" in data:
        wsheet.cell(row=i, column=25).value = 1
    else:
        wsheet.cell(row=i, column=25).value = 0

    data = rsheet.cell(row=i, column=9).value
    if data == "1 (Lowest)":
        wsheet.cell(row=i, column=26).value = 1
    elif data == 2:
        wsheet.cell(row=i, column=26).value = 2
    elif data == 3:
        wsheet.cell(row=i, column=26).value = 3
    elif data == 4:
        wsheet.cell(row=i, column=26).value = 4
    else:
        wsheet.cell(row=i, column=26).value = 5

print(wsheet.cell(row=i, column=23))

i += 1
while i < 205000:
   wsheet.cell(row=i, column=1).value = i - 1
   for j in range(2, 26):
       if j != 23:
           wsheet.cell(row=i, column=j).value = random.randint(0,1)
       else:
           wsheet.cell(row=i, column=j).value = random.randint(0,5)
   
   if wsheet.cell(row=i, column=2).value == 1 and wsheet.cell(row=i, column=23).value > 3 \
       and wsheet.cell(row=i, column=7).value == 1 and wsheet.cell(row=i, column=24).value == 1:
       wsheet.cell(row=i, column=26).value = 5
   elif (wsheet.cell(row=i, column=2).value == 1 and wsheet.cell(row=i, column=23).value > 3 \
         and wsheet.cell(row=i, column=7).value == 1)\
       or (wsheet.cell(row=i, column=24).value == 1 and wsheet.cell(row=i, column=7).value == 1) \
        or ((wsheet.cell(row=i, column=11).value == 1 or wsheet.cell(row=i, column=12).value == 1 or \
            wsheet.cell(row=i, column=13).value == 1 or wsheet.cell(row=i, column=14).value == 1 or \
                wsheet.cell(row=i, column=15).value == 1) and wsheet.cell(row=i, column=17).value == 1):
       wsheet.cell(row=i, column=26).value = 4
   elif (wsheet.cell(row=i, column=2).value == 1 and wsheet.cell(row=i, column=23).value >= 3) \
       or ((wsheet.cell(row=i, column=11).value == 1 or wsheet.cell(row=i, column=12).value == 1 \
            or wsheet.cell(row=i, column=13).value == 1 or wsheet.cell(row=i, column=14).value == 1 \
               or wsheet.cell(row=i, column=15).value == 1) and \
                   (wsheet.cell(row=i, column=9).value == 1 or wsheet.cell(row=i, column=7).value == 1 \
                    or wsheet.cell(row=i, column=8).value == 1)) or wsheet.cell(row=i, column=25).value == 1:
       wsheet.cell(row=i, column=26).value = 3
   else:
       wsheet.cell(row=i, column=26).value = random.randint(1,2)    
   i += 1


wbook.save("output.xlsx")
read_file = pd.read_excel ("output.xlsx")
read_file.to_csv ("output.csv", 
                  index = None,
                  header=True)