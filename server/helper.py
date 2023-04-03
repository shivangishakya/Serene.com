import openpyxl

stress_1 = ["You are good to go. Few Suggestions will be: Eat well, get enough sleep, be physically active, cut down \
    on alcohol, and take time to relax as well as working and studying. \
        Read about the 5 steps to mental wellbeing. Avoid drugs, including\
              lots of caffeine – this can have a negative impact on your \
                stress levels and wellbeing.",  "You can refer some videos:",
                    "https://www.youtube.com/watch?v=Bk2-dKH2Ta4",
                    "https://www.youtube.com/watch?v=eGVWRvNe1-A",
                    "https://www.youtube.com/watch?v=SWpOCLP1hjw"]

stress_2 = ["Keep a Detailed Calendar - A calendar that includes class schedule, assignment due dates, study times, social events and anything that will take up your time is a must. Start each semester by pulling class syllabi and adding in important dates to your digital calendar — or paper if you still prefer the written word. Color coding by class, due date and importance can help your organize. Set up automatic reminders on your phone if you're prone to procrastinate or forget dates.\
    Prioritize - When you add an item to your planner or digital calendar, color-code it with a “must do” or a “want to do” color. This will help you set priorities and make choices about how to spend your time. Make sure to build some flextime into your schedule so you have time for some fun opportunities when they come up. \
        Schedule Backwards - College has more long-term assignments than high school. Even if that paper isn’t due for another six weeks, get it on your calendar now and work backward from the due date. Set smaller, self-imposed due dates along the way to have parts of a large project complete so that you don’t have to pull the stressful all-nighter.\
Start a Routine - You set your own schedule in college, and it’s important to get in a groove. Figure out some basics as early as possible. How early do you want your first class? Where and when are you best able to focus studying? Where will you eat lunch and dinner? The fewer questions you have about these, the less stress you will feel when midterms and finals roll around. \
Exercise - Unless you’re on a college sports team, this one might not become an immediate priority. Even high school athletes can quickly let exercise fall to the wayside — but that’s a mistake. Join an intramural sports team, check out a group exercise class at the student gym or find a running buddy. Exercise will improve your mood and ward off the Freshman 15. Genius Tip: SignUpGenius can help you organize a walking group or a running club. \
Say NO – In high school, participating in a ton of activities might have been standard to beef up your resume, but you need to be more focused in college. Dip your toe into several groups if you’re not sure what you want to do, but hone in on your passions and choose a couple that mean the most to you. You may also be tempted to overload your schedule with parties and social events, but don’t let your new friends dictate how you spend your time. \
Know Your Sleep Requirements - Remember when your parents used to set your bedtime? Those days are over. Ideally, you’ll get eight hours of sleep each night, but it’s not just about quantity. If you get eight hours of sleep but don’t go to bed until 2 a.m. each night, your body is bound to feel out of whack.", 
"You can refer some videos:",
                    "https://www.youtube.com/watch?v=Bk2-dKH2Ta4",
                    "https://www.youtube.com/watch?v=eGVWRvNe1-A",
                    "https://www.youtube.com/watch?v=SWpOCLP1hjw"]

def stress_desc(val):
    if val == 1 or val == 2:
        return stress_1
    else:
        return stress_2
    
def format_survey(jsondata):
    ques = ['Primary Reasons of Stress', 'Have you been bullied?', \
        'Are you an International Student?', 'What do you miss the most about your home?',\
                'How many courses you have taken?', 'Financing the education', \
                'Is your commute to college stressful?']
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(ques)
    ws.append(jsondata)

    wb.save('quiz.xlsx')
    rbook = openpyxl.load_workbook('quiz.xlsx', data_only=True)
    rsheet = rbook.active

    responses = {}
    data = rsheet.cell(row=2, column=1).value
    if "Academic" in data:
        responses["Academics"] = 1
    else:
        responses["Academics"] = 0
    if "fitness" in data:
        responses["Looks_Fitness"] = 1
    else:
        responses["Looks_Fitness"] = 0
    if "Social" in data:
        responses["Social_life"] = 1
    else:
        responses["Social_life"] = 0
    if "extra-curricular" in data:
        responses["Xtra_curricular"] = 1
    else:
        responses["Xtra_curricular"] = 0
    if "Athletic" in data:
        responses["Athletics"] = 1
    else:
        responses["Athletics"] = 0
    if "Career" in data:
        responses["Career"] = 1
    else:
        responses["Career"] = 0
    if "Financial" in data:
        responses["Finance"] = 1
    else:
        responses["Finance"] = 0
    if "Relationship" in data:
        responses["Relationship"] = 1
    else:
        responses["Relationship"] = 0
    if "Cultural" in data:
        responses["Cultural_Shock"] = 1
    else:
        responses["Cultural_Shock"] = 0

    data = rsheet.cell(row=2, column=2).value
    if "Mentally/emotionally bullied" in data:
        responses["Emotional_bullied"] = 1
    else:
        responses["Emotional_bullied"] = 0
    if "Physically bullied" in data:
        responses["Physical_bullied"] = 1
    else:
        responses["Physical_bullied"] = 0
    if "Verbal bullying" in data:
        responses["Verbal_bullied"] = 1
    else:
        responses["Verbal_bullied"] = 0
    if "Social bullying" in data:
        responses["Social_bullied"] = 1
    else:
        responses["Social_bullied"] = 0
    if "Cyber bullying" in data:
        responses["Cyber_bullied"] = 1
    else:
        responses["Cyber_bullied"] = 0

    data = rsheet.cell(row=2, column=3).value
    if "Yes" in data:
        responses["International"] = 1
    else:
        responses["International"] = 0

    data = rsheet.cell(row=2, column=4).value
    if "No, I don't miss my home" in data:
        responses["Miss_home"] = 0
    else:
        responses["Miss_home"] = 1
    if "Yes, Family and friends" in data:
        responses["Family_friends"] = 1
    else:
        responses["Family_friends"] = 0
    if "Yes, Food" in data:
        responses["Food"] = 1
    else:
        responses["Food"] = 0
    if "Yes, Sensory experience of staying in home" in data:
        responses["Sensory"] = 1
    else:
        responses["Sensory"] = 0
    if "Yes, Social life of my hometown" in data:
        responses["Miss_social"] = 1
    else:
        responses["Miss_social"] = 0
    if "Yes, Native language conversations" in data:
        responses["Native_language"] = 1
    else:
        responses["Native_language"] = 0
    
    data = rsheet.cell(row=2, column=5).value
    if data == 1:
        responses["Courses"] = 1
    elif data == 2:
        responses["Courses"] = 2
    elif data == 3:
        responses["Courses"] = 3
    elif data == 4:
        responses["Courses"] = 4
    elif data == "More than 4":
        responses["Courses"] = 5
    else:
        responses["Courses"] = 0

    data = rsheet.cell(row=2, column=6).value
    if "Loan" in data:
        responses["Loan"] = 1
    else:
        responses["Loan"] = 0

    data = rsheet.cell(row=2, column=7).value
    if "Yes" in data or "Maybe" in data:
        responses["Stressed_commute"] = 1
    else:
        responses["Stressed_commute"] = 0
    
    return responses

    
def create_data(data):
    res = []
    list_cols = ['Academics', 'Looks_Fitness', 'Social_life', \
                 'Xtra_curricular', 'Athletics', 'Career', \
                    'Finance', 'Relationship', 'Cultural_Shock',\
                    'Emotional_bullied', 'Physical_bullied',\
                    'Verbal_bullied', 'Social_bullied',\
                    'Cyber_bullied', 'International',\
                    'Miss_home', 'Family_friends',\
                    'Food', 'Sensory', \
                    'Miss_social', \
                    'Native_language', 'Courses', 'Loan', \
                    'Stressed_commute']
    for i in list_cols:
        if i == 'Courses':
            if data['Courses'] == 3:
                res.append(1)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
            elif data['Courses'] == 2:
                res.append(0)
                res.append(1)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
            elif data['Courses'] == 4:
                res.append(0)
                res.append(0)
                res.append(1)
                res.append(0)
                res.append(0)
                res.append(0)
            elif data['Courses'] == 0:
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(1)
                res.append(0)
                res.append(0)
            elif data['Courses'] == 5:
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(1)
                res.append(0)
            elif data['Courses'] == 1:
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(1)
                res.append(0)
            else:
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
                res.append(0)
        else:
            if data[i] == 1:
                res.append(1)
                res.append(0)
            else:
                res.append(0)
                res.append(1)
    return res
